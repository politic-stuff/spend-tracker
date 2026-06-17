#!/usr/bin/env python3
# Import AdMo "Spending Chart" .xlsx export(s) into the tracker.
#   races[rk].spenders[] (advertiser + side + grand-total spend)
#   races[rk].buys[]     (advertiser x market x media-week x spend)
# Media weeks: Tue-Mon, reverse-counted from the Nov 3 general (W1 = week before).
# Usage:
#   python3 scripts/import-admo.py <file.xlsx> [more.xlsx ...]
#   python3 scripts/import-admo.py            # imports all ~/Downloads/political-spending-chart*.xlsx
import sys, json, os, re, glob, subprocess
from datetime import date, timedelta
import openpyxl

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, '..', 'data', 'data.json')
GENERAL = date(2026, 11, 3)

def parse_date(s):
    s = s or ''
    m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m: return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m: return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    return None

def tue_on_or_before(d): return d - timedelta(days=(d.weekday() - 1) % 7)
def media_week(fs):
    if not fs: return None
    return round(((tue_on_or_before(GENERAL) - timedelta(days=7)) - tue_on_or_before(fs)).days / 7) + 1
def side_of(p):
    p = (p or '').lower(); return 'D' if 'democrat' in p else 'R' if 'republican' in p else 'I' if p else None

STATES = set('AL AK AZ AR CA CO CT DE FL GA HI ID IL IN IA KS KY LA ME MD MA MI MN MS MO MT NE NV NH NJ NM NY NC ND OH OK OR PA RI SC SD TN TX UT VT VA WA WV WI WY DC'.split())

def one_race_key(name, races):
    n = (name or '').strip()
    m = re.match(r'^([A-Z]{2})\s+CD-?0*(\d+)\s+(\d{4})$', n)
    if m: rk = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    elif re.match(r'^[A-Z]{2}\s+Senate\s+\d{4}$', n): rk = re.sub(r'^([A-Z]{2})\s+Senate\s+(\d{4})$', r'\1-Sen-\2', n)
    elif re.match(r'^[A-Z]{2}\s+Governor\s+\d{4}$', n): rk = re.sub(r'^([A-Z]{2})\s+Governor\s+(\d{4})$', r'\1-Gov-\2', n)
    else: rk = None
    return rk if rk in races else None

def market_state(market):
    codes = [c for c in re.findall(r'\b([A-Z]{2})\b', market or '') if c in STATES]
    return codes  # may be 0 (national/digital), 1, or 2 (border DMA)

def hint_tokens(rk):
    st, mid = rk.split('-')[0], rk.split('-')[1]
    s = st.lower()
    if mid.isdigit():
        n, n0 = mid, mid.zfill(2)
        return [f"{s} cd-{n}", f"{s} cd-{n0}", f"{s} cd{n}", f"{s}-{n}", f"{s}-{n0}", f"cd-{n0}", f"cd-{n} "]
    if mid == 'Sen': return [f"{s} senate", f"{s} sen "]
    if mid == 'Gov': return [f"{s} governor", f"{s} gov "]
    return []

def attribute(adv, market, state_map, hints):
    codes = [c for c in market_state(market) if c in state_map]
    if len(codes) == 1: return state_map[codes[0]]
    a = ' ' + (adv or '').lower() + ' '
    for rk, toks in hints.items():
        if any(t in a for t in toks): return rk
    if len(codes) >= 1: return state_map[codes[0]]  # border DMA: take first matching state
    return None

def process_file(path, data):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    field = next((str(r[0]).replace('Race:', '').strip() for r in rows[:10] if r[0] and str(r[0]).startswith('Race:')), '')
    names = [x.strip() for x in field.split(',') if x.strip()]
    our = [(nm, one_race_key(nm, data['races'])) for nm in names]
    our = [(nm, rk) for nm, rk in our if rk]
    if not our: return [('SKIP', field[:55], 0, 0, 0)]
    multi = len(names) > 1
    state_map = {rk.split('-')[0]: rk for nm, rk in our}
    hints = {rk: hint_tokens(rk) for nm, rk in our}
    hdr_i = next(i for i, r in enumerate(rows) if r and r[0] == 'Party')
    hdr = rows[hdr_i]
    week_cols = []
    for c in range(3, len(hdr)):
        lab = str(hdr[c] or '')
        if ' - ' in lab and '/' in lab:
            s, e = lab.split(' - '); week_cols.append((c, parse_date(s), parse_date(e)))
    rd = {rk: {'sp': {}, 'buys': []} for nm, rk in our}
    cur_party = cur_adv = None
    for r in rows[hdr_i + 2:]:
        if r[0]: cur_party = r[0]
        if r[1]: cur_adv = r[1]
        market = r[2]
        if not cur_adv or not market: continue
        if 'total' in str(market).lower() or 'grand' in str(market).lower(): continue
        rk = our[0][1] if not multi else attribute(cur_adv, str(market), state_map, hints)
        if not rk: continue
        side = side_of(cur_party)
        gt = r[3] if isinstance(r[3], (int, float)) else 0
        rd[rk]['sp'].setdefault(cur_adv, {'side': side, 'amount': 0})['amount'] += gt or 0
        for (c, s, e) in week_cols:
            v = r[c] if c < len(r) and isinstance(r[c], (int, float)) else 0
            if v and v > 0:
                rd[rk]['buys'].append({'advertiser': cur_adv, 'side': side, 'market': str(market),
                    'flightStart': s.isoformat(), 'flightEnd': e.isoformat() if e else None,
                    'amount': round(v), 'week': media_week(s), 'source': 'AdImpact (AdMo)'})
    results = []
    for nm, rk in our:
        d = rd[rk]
        if not d['buys'] and not d['sp']: continue
        data['races'][rk]['spenders'] = [{'name': k, 'side': v['side'], 'amount': round(v['amount'])}
                                         for k, v in sorted(d['sp'].items(), key=lambda x: -x[1]['amount'])]
        data['races'][rk]['buys'] = d['buys']
        dem = sum(v['amount'] for v in d['sp'].values() if v['side'] == 'D')
        rep = sum(v['amount'] for v in d['sp'].values() if v['side'] == 'R')
        tot = sum(v['amount'] for v in d['sp'].values())
        data['races'][rk]['adimpact'] = {'demSide': round(dem), 'repSide': round(rep), 'total': round(tot), 'asOf': date.today().isoformat(), 'source': 'AdMo'}
        results.append((rk, nm, len(d['sp']), len(d['buys']), len(set(b['market'] for b in d['buys']))))
    return results

files = sys.argv[1:] or sorted(glob.glob(os.path.expanduser('~/Downloads/political-spending-chart*.xlsx')))
if not files: print('No xlsx files given or found in ~/Downloads.'); sys.exit(1)
data = json.load(open(DATA))
done = []
for f in files:
    print(f"{os.path.basename(f)}:")
    try:
        for res in process_file(f, data):
            done.append(res); print(f"  {res[0]:14} {res[1]}: {res[2]} spenders, {res[3]} buys, {res[4]} markets")
    except Exception as e:
        print(f"  ERROR: {e}")
ok = [d for d in done if d[0] != 'SKIP']
data['lastUpdated'] = date.today().isoformat()
data.setdefault('changelog', []).insert(0, {'ts': date.today().isoformat(),
    'note': f"AdMo import: {len(ok)} race(s) backfilled — " + ', '.join(d[0] for d in ok[:8])})
data['changelog'] = data['changelog'][:40]
json.dump(data, open(DATA, 'w'), indent=2)
subprocess.run(['node', os.path.join(HERE, 'wrap.js')])
print(f"\nDone: {len(ok)} race(s) imported, {len([d for d in done if d[0]=='SKIP'])} skipped (not a tracked race).")
